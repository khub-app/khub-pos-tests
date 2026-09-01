"""Reads data/test_data.xlsx's RunConfig sheet for run_tests.ps1.

Mirrors khub-web-tests' scripts/read_run_config.py, scaled down to this
project's actual setup: Appium drives ONE device (emulator or real_device)
at a time, so there's no Headless/Workers/multi-environment toggle here -
just which environment to target and which suite(s) (smoke/regression) to
run, both matching what conftest.py's Excel-driven markers already expect.
"""
import json
import os
import sys

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_XLSX = os.path.join(os.path.dirname(_SCRIPT_DIR), "data", "test_data.xlsx")

_DEFAULT_ENVIRONMENT = "emulator"


def read_config():
    try:
        import openpyxl
    except ImportError:
        print(json.dumps({"error": "openpyxl not installed"}))
        sys.exit(1)

    if not os.path.exists(_XLSX):
        print(json.dumps({"error": f"Excel file not found: {_XLSX}"}))
        sys.exit(1)

    wb = openpyxl.load_workbook(_XLSX, read_only=True, data_only=True)

    if "RunConfig" not in wb.sheetnames:
        print(json.dumps({"error": "RunConfig sheet not found in data/test_data.xlsx"}))
        sys.exit(1)

    ws = wb["RunConfig"]
    environment = _DEFAULT_ENVIRONMENT
    suites = {}
    section = None  # "suite" once we're past the "Suite | Enabled" header

    for row in ws.iter_rows(values_only=True):
        if not row or row[0] is None:
            continue  # blank row - stay in whatever section we're in
        key = str(row[0]).strip()
        raw = row[1] if len(row) > 1 else None
        val = str(raw).strip().lower() if raw is not None else ""

        if key.lower() in ("setting", "value"):
            continue
        if key.lower() == "suite":
            section = "suite"
            continue
        if key.lower() == "enabled":
            continue

        if key.lower() == "environment":
            environment = str(raw).strip() if raw else _DEFAULT_ENVIRONMENT
        elif section == "suite":
            suites[key.lower()] = val in ("yes", "true", "y", "1")

    return {
        "environment": environment,
        "suites": [s for s, enabled in suites.items() if enabled],
    }


if __name__ == "__main__":
    print(json.dumps(read_config()))
