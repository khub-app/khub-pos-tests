import os
import shutil
import subprocess

import openpyxl
import pytest

from utilities.driver_manager import DriverManager
from utilities.logger import get_logger
from utilities.report_metadata import write_environment_properties, write_executor_json
from utilities.screenshot_helper import take_screenshot
from utilities.video_helper import start_recording, stop_recording_and_save

logger = get_logger(__name__)


def pytest_addoption(parser):
    parser.addoption(
        "--env",
        action="store",
        default=None,
        help="Target environment defined in config/config.yaml (e.g. emulator, real_device)",
    )


# ------------------ Dynamic markers from Excel Tags sheet ------------------
# Mirrors khub-web-tests' conftest.py: data/test_data.xlsx's "Tags" sheet has
# one row per test file (FilePath | smoke | regression | Date Added). A "Yes"
# in a marker column auto-applies that pytest marker to every item collected
# from that file, so `pytest -m smoke` / `-m regression` are driven by the
# spreadsheet instead of hardcoded decorators. New test file -> add a row
# here (exact relative path, matched as a path suffix) or it gets no markers.

def _load_tags_from_excel():
    """Read the Tags sheet and return {filepath: [marker_names]}."""
    excel_path = os.path.join(os.path.dirname(__file__), "data", "test_data.xlsx")
    if not os.path.exists(excel_path):
        return {}
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    if "Tags" not in wb.sheetnames:
        wb.close()
        return {}
    ws = wb["Tags"]
    headers = [cell.value for cell in ws[1]]
    marker_names = headers[1:]  # skip FilePath column
    tag_map = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        filepath = row[0]
        if not filepath:
            continue
        filepath = filepath.replace("\\", "/")
        markers = []
        for i, val in enumerate(row[1:]):
            if val and str(val).strip().lower() == "yes" and i < len(marker_names):
                markers.append(marker_names[i])
        if markers:
            tag_map[filepath] = markers
    wb.close()
    return tag_map


def pytest_collection_modifyitems(config, items):
    """Auto-apply markers from the Excel Tags sheet to collected test items."""
    tag_map = _load_tags_from_excel()
    if not tag_map:
        return
    applied_count = 0
    for item in items:
        item_path = str(item.fspath).replace("\\", "/")
        for tagged_path, markers in tag_map.items():
            if item_path.endswith(tagged_path):
                for marker_name in markers:
                    item.add_marker(getattr(pytest.mark, marker_name))
                applied_count += 1
                break
    if applied_count:
        print(f"\n[INFO] Applied Excel tags to {applied_count} test items from {len(tag_map)} tagged files")


# Holds whatever driver is currently active so the _step_media fixture below
# can find it even though the `driver` fixture itself is class-scoped (a
# function-scoped fixture in the same class gets a *different* request.node
# than the class-scoped one, so it can't rely on request to reach it).
_active_driver = {"instance": None}


@pytest.fixture(scope="class")
def driver(request):
    env_name = request.config.getoption("--env")
    manager = DriverManager(env_name)
    manager.reset_app_data()  # every class starts logged out, regardless of run order
    driver_instance = manager.start_driver()  # app renders landscape by default, no forcing needed
    _active_driver["instance"] = driver_instance
    yield driver_instance
    _active_driver["instance"] = None
    manager.stop_driver()


# Per-step screen recording (see _step_media below) - OFF by default. Each
# start/stop is a genuine adb screenrecord round-trip and measurably slowed
# the full 51-step suite down (confirmed live). Screenshots stay on
# regardless - they're cheap. Flip to "1" (env var, no code change needed)
# to turn video back on for a specific run.
_RECORD_VIDEO = os.environ.get("POS_RECORD_VIDEO", "0") == "1"


@pytest.fixture(autouse=True)
def _step_media(request):
    """Attaches a screenshot for this test method (one logical UI step)
    right after it finishes - for every step, pass or fail, not just
    failures. Mirrors khub-web-tests' per-test screenshot capture.

    Optionally also records a screen video per step (see _RECORD_VIDEO) -
    PER STEP rather than per whole scenario, because Android's screenrecord
    has a hard 3-minute cap per recording and a full multi-step scenario
    can run several minutes end to end, while no single step does. Videos
    are saved to videos/ on disk but NOT attached to the Allure report
    itself (only a text pointer to the file is) - embedding dozens of
    per-step video clips would bloat the --single-file HTML report exactly
    the way khub-web-tests' conftest.py explicitly avoids doing with its
    Playwright trace files.
    """
    driver_instance = _active_driver["instance"]
    if driver_instance is not None and _RECORD_VIDEO:
        start_recording(driver_instance)

    yield

    if driver_instance is None:
        return

    video_path = stop_recording_and_save(driver_instance, request.node.name) if _RECORD_VIDEO else None

    rep_call = getattr(request.node, "rep_call", None)
    failed = bool(rep_call and rep_call.failed)

    try:
        import allure

        try:
            screenshot_path = take_screenshot(driver_instance, request.node.name)
            with open(screenshot_path, "rb") as image_file:
                allure.attach(
                    image_file.read(),
                    name="Failure Screenshot" if failed else "Screenshot",
                    attachment_type=allure.attachment_type.PNG,
                )
        except Exception as e:
            logger.warning(f"Could not capture step screenshot: {e}")

        if video_path:
            allure.attach(
                f"Saved to {video_path}",
                name="Screen Recording",
                attachment_type=allure.attachment_type.TEXT,
            )
    except ImportError:
        pass


# ------------------ Skip later steps after an earlier one fails ------------------
# Several scenario test files (test_split_payment.py, test_sale_return.py, ...)
# are split into one pytest test method per logical UI step, sharing state via
# class attributes, so each step gets its own Allure page instead of being a
# flat allure.step() list inside one monolithic test. Those steps depend on
# each other's app state (e.g. "add card payment" only makes sense if "add
# cash payment" actually succeeded), so a failed step must stop the rest of
# that class's steps from running against unknown state, while still letting
# UNRELATED classes run normally.
#
# pytest-depends was tried for this and rejected: it works, but its
# `pytest_collection_modifyitems` hook unconditionally topologically-sorts
# the ENTIRE session's item list (confirmed by reading its source), which
# interleaves independent classes' steps breadth-first (all classes' step 1,
# then all classes' step 2, ...) instead of preserving each class's natural
# run - and since this app's class-scoped `driver` fixture only stays alive
# for a CONTIGUOUS run of one class's tests, that interleaving forced the
# fixture to tear down and recreate (full app reset + re-login) between
# every single step instead of once per class, corrupting the very state
# these steps were sharing. This plain hook-based approach instead only
# ever SKIPS - it never reorders anything, so pytest's own natural
# file -> class -> method definition order is left untouched.
_class_failed_step = {}


def pytest_runtest_setup(item):
    cls = item.cls
    if cls is not None and cls in _class_failed_step:
        pytest.skip(f"Skipped: an earlier step in this class failed ({_class_failed_step[cls]})")


@pytest.hookimpl(tryfirst=True, hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    report = outcome.get_result()

    # Stash the report on the item (standard pytest recipe) so the
    # _step_media fixture's teardown code - which runs during the
    # "teardown" phase, after this hook has already produced the "call"
    # phase's report - can read whether the step passed or failed.
    setattr(item, f"rep_{report.when}", report)

    if report.when == "call" and report.failed:
        if item.cls is not None:
            _class_failed_step[item.cls] = item.name
        logger.error(f"Test failed: {item.name}")


# ------------------ Run-wide pass/fail counters for environment.properties ------------------
_run_stats = {"passed": 0, "failed": 0, "skipped": 0, "errors": 0}


def pytest_runtest_logreport(report):
    if report.when == "setup" and report.outcome == "failed":
        _run_stats["errors"] += 1
    elif report.when == "call":
        if report.outcome == "passed":
            _run_stats["passed"] += 1
        elif report.outcome == "failed":
            _run_stats["failed"] += 1
    if report.outcome == "skipped" and report.when in ("setup", "call"):
        _run_stats["skipped"] += 1


def pytest_sessionfinish(session, exitstatus):
    """Auto-generates the Allure HTML report from this run's results, so a
    plain `pytest` invocation is enough - no separate manual
    `allure generate` step (mirrors khub-web-tests' conftest.py, scaled down
    to this project's single-env setup - no per-env report dirs/archiving).

    pytest.ini's addopts passes --alluredir=reports/allure-results
    --clean-alluredir, so every `pytest` invocation starts that raw-results
    folder empty - without it, results from every run all session long
    accumulated together (confirmed live: a single-test run's report ended
    up mixing in failures from unrelated tests run hours earlier). This
    generate step's own --clean flag is a second, independent safeguard on
    the HTML output directory itself.

    --single-file bundles everything (JS/CSS/JSON data) inline into one
    index.html instead of Allure's default multi-file layout (a data/
    folder with one JSON per test case, plus separate widgets/history/
    plugin/export folders and app.js/styles.css) - keeps reports/allure-report/
    to just that one file rather than 200+ scattered ones."""
    results_dir = session.config.getoption("--alluredir") or "reports/allure-results"
    if not os.path.isdir(results_dir) or not os.listdir(results_dir):
        return

    total = sum(_run_stats.values())
    stats = {"total": total, **_run_stats}
    try:
        write_environment_properties(results_dir, stats, session.config.getoption("--env"))
        write_executor_json(results_dir)
    except Exception as e:
        print(f"[WARNING] Could not write Allure environment/executor metadata: {e}")

    allure_cmd = shutil.which("allure")
    if not allure_cmd:
        print("\n[WARNING] Allure CLI not found on PATH - skipping report generation.")
        print(f"  To generate manually: allure generate {results_dir} -o reports/allure-report --clean --single-file")
        return

    report_dir = os.path.join(os.path.dirname(os.path.normpath(results_dir)), "allure-report")
    print("\n[INFO] Generating Allure HTML report...")
    try:
        result = subprocess.run(
            [allure_cmd, "generate", results_dir, "-o", report_dir, "--clean", "--single-file"],
            capture_output=True, text=True, timeout=120,
        )
    except Exception as e:
        print(f"[WARNING] Allure generate failed to run: {e}")
        return

    if result.returncode == 0:
        print(f"[INFO] Allure report: {os.path.abspath(os.path.join(report_dir, 'index.html'))}")
        print("  (open directly, or run 'allure serve reports/allure-results' for a live view)")
    else:
        print(f"[WARNING] Allure generate failed (exit {result.returncode})")
        if result.stderr:
            print(f"  {result.stderr.strip()[:300]}")
